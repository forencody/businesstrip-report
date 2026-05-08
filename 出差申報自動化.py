#!/usr/bin/env python3
"""
出差申報自動化工具
從 Google 行事曆自動產生差旅報支申請單

使用方式：
  python 出差申報自動化.py            → 自動處理上個月
  python 出差申報自動化.py 2026 3     → 處理 2026 年 3 月
"""

import json
import os
import pickle
import sys
import urllib.parse
import urllib.request
from calendar import monthrange
from copy import copy as _copy_obj
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

try:
    import anthropic as _anthropic
    _ANTHROPIC_AVAILABLE = True
except ImportError:
    _ANTHROPIC_AVAILABLE = False

# Google Calendar API
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

# ────────────────────────────────────────────
SCOPES = ["https://www.googleapis.com/auth/calendar.readonly"]
BASE_DIR = Path(__file__).parent
CONFIG_FILE = BASE_DIR / "destinations.json"
TEMPLATE_FILE = BASE_DIR / "陳峙霖2026差旅報支申請單(油費連動)2月.xlsx"
TOKEN_FILE = BASE_DIR / "token.pickle"
CREDENTIALS_FILE = BASE_DIR / "credentials.json"

# 油費報告書資料從第幾列開始（固定結構）
OIL_DATA_START_ROW = 4
OIL_DATA_MAX_ROWS = 26   # rows 4–29

ARROW_SYMBOL = "ßà"


# ════════════════════════════════════════════
# 設定讀取
# ════════════════════════════════════════════

def load_config() -> dict:
    if not CONFIG_FILE.exists():
        sys.exit(f"❌ 找不到設定檔：{CONFIG_FILE}\n"
                 "請確認 destinations.json 與本程式在同一資料夾")
    with open(CONFIG_FILE, encoding="utf-8") as f:
        return json.load(f)


# ════════════════════════════════════════════
# Google Calendar 認證
# ════════════════════════════════════════════

def get_calendar_service():
    # CI 環境：從環境變數還原 token.pickle 和 credentials.json
    token_b64 = os.environ.get("TOKEN_PICKLE_B64")
    if token_b64 and not TOKEN_FILE.exists():
        import base64
        TOKEN_FILE.write_bytes(base64.b64decode(token_b64))

    creds_json = os.environ.get("CREDENTIALS_JSON")
    if creds_json and not CREDENTIALS_FILE.exists():
        CREDENTIALS_FILE.write_text(creds_json, encoding="utf-8")

    if not CREDENTIALS_FILE.exists():
        sys.exit(
            f"❌ 找不到 Google API 憑證：{CREDENTIALS_FILE}\n"
            "請依照 SETUP.md 的說明下載並放置 credentials.json"
        )

    creds = None
    if TOKEN_FILE.exists():
        with open(TOKEN_FILE, "rb") as f:
            creds = pickle.load(f)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                str(CREDENTIALS_FILE), SCOPES
            )
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, "wb") as f:
            pickle.dump(creds, f)

    return build("calendar", "v3", credentials=creds)


# ════════════════════════════════════════════
# 行事曆搜尋
# ════════════════════════════════════════════

def fetch_events(service, config: dict, year: int, month: int) -> list:
    keyword = config["settings"]["calendar_keyword"]
    _, last_day = monthrange(year, month)

    time_min = f"{year}-{month:02d}-01T00:00:00+08:00"
    time_max = f"{year}-{month:02d}-{last_day:02d}T23:59:59+08:00"

    # 取得所有日曆清單
    calendars = service.calendarList().list().execute().get("items", [])

    all_events = []
    seen_ids = set()
    for cal in calendars:
        cal_id = cal["id"]
        try:
            result = service.events().list(
                calendarId=cal_id,
                timeMin=time_min,
                timeMax=time_max,
                q=keyword,
                singleEvents=True,
                orderBy="startTime",
                maxResults=200,
            ).execute()
            for ev in result.get("items", []):
                if ev["id"] not in seen_ids:
                    seen_ids.add(ev["id"])
                    all_events.append(ev)
        except Exception:
            pass  # 略過無權限的日曆

    # 依開始時間排序
    def sort_key(ev):
        s = ev.get("start", {})
        return s.get("date") or s.get("dateTime", "")

    all_events.sort(key=sort_key)
    return all_events


# ════════════════════════════════════════════
# 事件解析
# ════════════════════════════════════════════

def _parse_date(event: dict) -> tuple[date, date]:
    """回傳 (start_date, end_date)，end 為最後一天（含）"""
    start_raw = event.get("start", {})
    end_raw = event.get("end", {})

    if "date" in start_raw:                         # 全天事件
        start = date.fromisoformat(start_raw["date"])
        end = date.fromisoformat(end_raw["date"])
        end = date(end.year, end.month, end.day - 1) if end > start else start
    else:                                           # 時段事件
        from datetime import datetime, timezone
        start = datetime.fromisoformat(
            start_raw["dateTime"].replace("Z", "+00:00")
        ).astimezone().date()
        end = start

    return start, end


def _event_duration_days(start: date, end: date) -> int:
    return (end - start).days + 1


def _ai_lookup_district(company_name: str, destinations: dict) -> str:
    """
    用 Claude API 查詢公司在台灣的所在地，回傳 destinations 中最接近的 key。
    找不到或 API 不可用時回傳空字串。
    """
    if not _ANTHROPIC_AVAILABLE:
        return ""
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return ""
    try:
        client = _anthropic.Anthropic(api_key=api_key)
        district_list = "、".join(destinations.keys())
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=30,
            messages=[{
                "role": "user",
                "content": (
                    f"台灣公司「{company_name}」的總部或主要廠區在哪個縣市或區域？\n"
                    f"請從以下選項中選出最符合的一個，只回答選項名稱，不要其他文字：\n"
                    f"{district_list}\n"
                    "若不確定請回答「不確定」。"
                ),
            }],
        )
        result = msg.content[0].text.strip()
        return result if result in destinations else ""
    except Exception:
        return ""


def _google_maps_distance_km(origin: str, destination_city: str, api_key: str) -> int:
    """
    用 Google Maps Routes API 查詢駕車距離（公里，四捨五入）。
    失敗時回傳 0。
    """
    if not api_key:
        return 0
    try:
        import urllib.request
        url = "https://routes.googleapis.com/directions/v2:computeRoutes"
        payload = json.dumps({
            "origin": {"address": origin},
            "destination": {"address": destination_city},
            "travelMode": "DRIVE",
            "routingPreference": "TRAFFIC_UNAWARE",
        }).encode("utf-8")
        req = urllib.request.Request(url, data=payload, method="POST", headers={
            "Content-Type": "application/json",
            "X-Goog-Api-Key": api_key,
            "X-Goog-FieldMask": "routes.distanceMeters",
        })
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
        meters = data["routes"][0]["distanceMeters"]
        return round(meters / 1000)
    except Exception:
        pass
    return 0


def _resolve_mileage(dest_info: dict, config: dict, district: str) -> int:
    """
    依 use_hsr 旗標決定里程：
    - use_hsr = true  → hsr_mileage（固定 33 km）
    - use_hsr = false → 查詢成創到地區城市的來回距離並快取
    """
    settings = config["settings"]
    if dest_info.get("use_hsr", True):
        return settings.get("hsr_mileage", 33)

    cached = dest_info.get("mileage", 0)
    if cached > 0:
        return cached

    api_key = settings.get("google_maps_api_key", "") or os.environ.get("GOOGLE_MAPS_API_KEY", "")
    if not api_key:
        return 0

    origin = settings.get("company_address", "台南市北區開元路457號")
    city = dest_info.get("city", district)
    one_way = _google_maps_distance_km(origin, city, api_key)
    if one_way > 0:
        km = one_way * 2
        config["destinations"][district]["mileage"] = km
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        print(f"   🗺  Google Maps：成創 ↔ {city} = {one_way} km × 2 = {km} km（已快取）")
        return km
    return 0


def _resolve_company_mileage(company_key: str, company_name: str,
                              dest_info: dict, config: dict, district: str) -> int:
    """
    use_hsr=false 且有匹配到公司時，用公司名稱查詢實際來回距離。
    優先讀 companies[key]["mileage"] 快取；
    若快取為 0，以 companies[key]["address"]（若有）或公司名稱查 Google Maps。
    找不到時退回地區級查詢。
    """
    settings = config["settings"]
    if dest_info.get("use_hsr", True):
        return settings.get("hsr_mileage", 33)

    company_info = config.get("companies", {}).get(company_key, {})
    cached = company_info.get("mileage", 0)
    if cached > 0:
        return cached

    api_key = settings.get("google_maps_api_key", "") or os.environ.get("GOOGLE_MAPS_API_KEY", "")
    if not api_key:
        return _resolve_mileage(dest_info, config, district)

    origin = settings.get("company_address", "台南市北區開元路457號")
    # 用 address 欄位（若有手動設定），否則用公司名稱讓 Google Maps 自行解析
    dest_query = company_info.get("address") or company_name
    one_way = _google_maps_distance_km(origin, dest_query, api_key)
    if one_way > 0:
        km = one_way * 2
        config["companies"][company_key]["mileage"] = km
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        print(f"   🗺  Google Maps：成創 ↔ {dest_query} = {one_way} km × 2 = {km} km（已快取）")
        return km

    # 退回地區級
    return _resolve_mileage(dest_info, config, district)


def _is_skippable(title: str, location: str, config: dict) -> str:
    """
    判斷事件是否應跳過（非實體出差）。
    回傳跳過原因字串，若不跳過回傳空字串。
    """
    filters = config.get("filters", {})
    online_kws = filters.get("online_keywords", ["線上", "Teams", "Zoom", "Meet"])
    internal_kws = filters.get("internal_location_keywords", ["會議室", "辦公室"])
    skip_types = filters.get("skip_event_types", ["請假"])

    # 線上會議
    for kw in online_kws:
        if kw in title or kw in location:
            return f"線上會議（含「{kw}」）"

    # 標記為請假/休假
    for kw in skip_types:
        if kw in title:
            return f"非出差事件（含「{kw}」）"

    # 斜線格式：取第二段（客戶/地點欄位）判斷是否為會議室
    parts = [p.strip() for p in title.split("/")]
    client_field = parts[1] if len(parts) > 1 else ""
    for kw in internal_kws:
        if kw in client_field:
            return f"內部地點（「{client_field}」含「{kw}」）"

    # 第二段是本人名字 → 內部會議（格式：歸屬/本人名/內容）
    person_name = config.get("settings", {}).get("person_name", "")
    if person_name and client_field == person_name:
        return f"內部會議（客戶欄為本人名「{person_name}」）"

    return ""


def parse_event(event: dict, config: dict):
    """
    依照行事曆標題格式「行程歸屬/客戶地點/與會人/會議內容」解析出差記錄。
    若判定為線上或內部會議則自動跳過（回傳 None）。
    """
    settings = config["settings"]
    companies = config.get("companies", {})
    destinations = config.get("destinations", {})
    work_reason_map = config.get("work_reason_map", {})

    title = event.get("summary", "") or ""
    location = event.get("location", "") or ""

    # ── 過濾：線上/內部/請假 ────────────────────
    skip_reason = _is_skippable(title, location, config)
    if skip_reason:
        return None   # 靜默跳過，不顯示在確認清單

    try:
        start_date, end_date = _parse_date(event)
    except Exception:
        return None

    days = _event_duration_days(start_date, end_date)

    # ── 解析斜線結構 ──────────────────────────
    # 格式：行程歸屬 / 客戶地點 / 與會人 / 會議內容
    parts = [p.strip() for p in title.split("/")]
    client_field = parts[1] if len(parts) > 1 else title
    content_field = parts[-1] if len(parts) > 1 else ""

    # ── 海外判斷：起迄點顯示桃園，里程固定 33 km ──
    overseas_kws = settings.get("overseas_keywords",
                                config.get("filters", {}).get("overseas_keywords", []))
    if not overseas_kws:
        overseas_kws = config.get("filters", {}).get("overseas_keywords", [])
    is_overseas = any(kw in title or kw in location for kw in overseas_kws)
    if is_overseas:
        work_reason = settings["default_work_reason"]
        for pattern, reason in config.get("work_reason_map", {}).items():
            if pattern in content_field or pattern in title:
                work_reason = reason
                break
        meal_allowance = settings["meal_outside_tainan"] * days
        return [{
            "month": start_date.month,
            "day": start_date.day,
            "end_day": end_date.day,
            "days": days,
            "destination": "桃園",
            "mileage": settings.get("hsr_mileage", 33),
            "work_reason": work_reason,
            "company": client_field or title,
            "inside_tainan": False,
            "meal_allowance": meal_allowance,
            "daily_meal": settings["meal_outside_tainan"],
            "_title": title,
            "_location": location,
        }]

    # ── 找公司及目的地 ──────────────────────────
    company_name = ""
    company_key = ""
    district = ""
    mileage = settings.get("hsr_mileage", 33)
    inside_tainan = False

    # 優先用 client_field（第二段）對照公司表
    for keyword, info in companies.items():
        if keyword in client_field:
            company_key = keyword
            company_name = info.get("full_name", client_field)
            district = info.get("district", "")
            break

    # 若沒對到，全文比對公司表
    if not company_name:
        for keyword, info in companies.items():
            if keyword in title:
                company_key = keyword
                company_name = info.get("full_name", keyword)
                district = info.get("district", "")
                break

    # 取得目的地里程資訊
    if district and district in destinations:
        dest_info = destinations[district]
        inside_tainan = dest_info.get("inside_tainan", False)
        if company_key and not dest_info.get("use_hsr", True):
            # use_hsr=false：用公司實際地址查來回距離
            mileage = _resolve_company_mileage(company_key, company_name, dest_info, config, district)
        else:
            mileage = _resolve_mileage(dest_info, config, district)
    else:
        # 嘗試從地名表直接匹配
        for dest_name, dest_info in destinations.items():
            if dest_name in title or dest_name in location:
                district = dest_name
                inside_tainan = dest_info.get("inside_tainan", False)
                mileage = _resolve_mileage(dest_info, config, district)
                break

    # 若公司名稱仍空，用 client_field 原文
    if not company_name:
        company_name = client_field

    # 仍找不到地點 → 用 AI 查詢
    if not district and company_name:
        ai_district = _ai_lookup_district(company_name, destinations)
        if ai_district:
            district = ai_district
            dest_info = destinations[district]
            inside_tainan = dest_info.get("inside_tainan", False)
            if company_key and not dest_info.get("use_hsr", True):
                mileage = _resolve_company_mileage(company_key, company_name, dest_info, config, district)
            else:
                mileage = _resolve_mileage(dest_info, config, district)
            print(f"   🤖 AI 查詢「{company_name}」→ {district}")

    # ── 工作事由 ──────────────────────────────
    work_reason = settings["default_work_reason"]
    for pattern, reason in work_reason_map.items():
        if pattern in content_field or pattern in title:
            work_reason = reason
            break

    # ── 膳雜費（依天數計算）───────────────────
    daily_meal = (
        settings["meal_inside_tainan"] if inside_tainan
        else settings["meal_outside_tainan"]
    )
    meal_allowance = daily_meal * days

    return [{
        "month": start_date.month,
        "day": start_date.day,
        "end_day": end_date.day,
        "days": days,
        "destination": district or "",
        "mileage": mileage,
        "work_reason": work_reason,
        "company": company_name or title,
        "inside_tainan": inside_tainan,
        "meal_allowance": meal_allowance,
        "daily_meal": daily_meal,
        "_title": title,
        "_location": location,
    }]


# ════════════════════════════════════════════
# 同日出差合併
# ════════════════════════════════════════════

def merge_same_day_trips(trips: list, config: dict) -> list:
    """
    同一天有多筆出差時：
    - 取里程最長的作為主要目的地
    - 公司名稱合併列入備註
    - 膳雜費只計一天
    """
    from collections import OrderedDict
    grouped = OrderedDict()
    for trip in trips:
        key = (trip["month"], trip["day"])
        grouped.setdefault(key, []).append(trip)

    merged = []
    for (month, day), group in grouped.items():
        if len(group) == 1:
            merged.append(group[0])
            continue

        # 找里程最長的作為主行程
        primary = max(group, key=lambda t: t["mileage"])
        companies = "、".join(
            t["company"] for t in group if t["company"]
        )
        # 膳雜費以主行程的 daily_meal × days 為準（一天只算一次）
        meal = primary["daily_meal"] * primary["days"]

        merged_trip = dict(primary)
        merged_trip["company"] = companies
        merged_trip["meal_allowance"] = meal

        destinations_str = "、".join(
            t["destination"] for t in group if t["destination"] and t["destination"] != primary["destination"]
        )
        if destinations_str:
            print(f"   🔀 {month:02d}/{day:02d} 同日合併：{primary['destination']}（最遠）+ {destinations_str}")
        else:
            print(f"   🔀 {month:02d}/{day:02d} 同日合併：{len(group)} 筆 → 目的地 {primary['destination']}")

        merged.append(merged_trip)

    return merged


# ════════════════════════════════════════════
# 互動確認 / 編輯
# ════════════════════════════════════════════

def _save_company_mapping(company: str, district: str, config: dict):
    """將新的公司→目的地對應存回 destinations.json，下次自動帶入"""
    if not company or not district:
        return
    dest_info = config.get("destinations", {}).get(district)
    if not dest_info:
        return  # 目的地不在清單內，不儲存
    config.setdefault("companies", {})[company] = {
        "district": district,
        "full_name": company,
    }
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    print(f"       💾 已記住：「{company}」→「{district}」，下次自動帶入")


def confirm_trips(trips: list, config: dict):
    print("\n" + "═" * 62)
    print("  找到以下出差記錄，請逐筆確認")
    print("═" * 62)

    confirmed = []
    for i, trip in enumerate(trips, 1):
        if trip["days"] > 1:
            end_d = trip.get("end_day", trip["day"] + trip["days"] - 1)
            date_str = f"{trip['month']:02d}/{trip['day']:02d}~{end_d:02d}（{trip['days']}天）"
        else:
            date_str = f"{trip['month']:02d}/{trip['day']:02d}"
        dest_display = trip["destination"] or "⚠ 未知目的地"
        print(f"\n[{i:2d}]  {date_str}  "
              f"{dest_display}  ─  {trip['company']}")
        print(f"       里程 {trip['mileage']} km  ｜  事由：{trip['work_reason']}  "
              f"｜  膳雜費：{trip['meal_allowance']} 元")
        print(f"       行事曆：{trip['_title']}"
              + (f"  @{trip['_location']}" if trip["_location"] else ""))

        if not trip["destination"]:
            print("       ⚠  無法辨識目的地，請輸入 e 編輯")

        action = input("  → [Enter=確認  n=跳過  e=編輯]: ").strip().lower()
        if action == "n":
            continue
        if action == "e":
            old_dest = trip["destination"]
            trip = _edit_trip(trip, config)
            # 若目的地有變動，詢問是否記住
            new_dest = trip["destination"]
            if new_dest and new_dest != old_dest and trip["company"]:
                save = input(f"       記住「{trip['company']}」→「{new_dest}」以後自動帶入？[y/n]: ").strip().lower()
                if save == "y":
                    _save_company_mapping(trip["company"], new_dest, config)
        confirmed.append(trip)

    return confirmed


def _edit_trip(trip: dict, config: dict) -> dict:
    destinations = config.get("destinations", {})

    def _prompt(label, current):
        val = input(f"     {label} [{current}]: ").strip()
        return val if val else current

    print("   ── 編輯（直接 Enter 保留原值）──")

    # 目的地：輸入後自動查里程
    new_dest = input(f"     目的地（迄點） [{trip['destination']}]: ").strip()
    if new_dest:
        trip["destination"] = new_dest
        if new_dest in destinations:
            dest_info = destinations[new_dest]
            trip["mileage"] = _resolve_mileage(dest_info, config, new_dest)
            trip["inside_tainan"] = dest_info.get("inside_tainan", False)
            daily = (config["settings"]["meal_inside_tainan"] if trip["inside_tainan"]
                     else config["settings"]["meal_outside_tainan"])
            trip["meal_allowance"] = daily * trip["days"]
            print(f"       → 自動帶入里程 {trip['mileage']} km，膳雜費 {trip['meal_allowance']} 元")

    trip["company"] = _prompt("備註（公司名）", trip["company"])
    trip["work_reason"] = _prompt("工作事由", trip["work_reason"])

    raw_m = input(f"     里程 [{trip['mileage']}]: ").strip()
    if raw_m:
        trip["mileage"] = int(raw_m)

    raw_meal = input(f"     膳雜費 [{trip['meal_allowance']}]: ").strip()
    if raw_meal:
        trip["meal_allowance"] = int(raw_meal)

    return trip


# ════════════════════════════════════════════
# Excel 產生
# ════════════════════════════════════════════

def _month_zh(month: int) -> str:
    return f"{month}月"


def _rename_sheet_and_update_refs(wb, old_name: str, new_name: str):
    """重新命名工作表，並更新其他工作表中對它的公式引用"""
    if old_name not in wb.sheetnames:
        return
    wb[old_name].title = new_name
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if (
                    cell.value
                    and isinstance(cell.value, str)
                    and cell.value.startswith("=")
                    and old_name in cell.value
                ):
                    cell.value = cell.value.replace(old_name + "!", new_name + "!")


def _find_total_row(ws, keyword="合計"):
    """找到合計列的列號（搜尋前 8 欄，含合併儲存格）"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == keyword and cell.column <= 8:
                return cell.row
    return None


def _copy_row_style(ws, source_row: int, target_row: int):
    """
    複製一整列的儲存格格式（框線、填色、字體、列高）到目標列。
    不複製合併儲存格——合併會讓新列的儲存格變成唯讀 MergedCell，
    導致後續寫入資料被 safe_set 跳過。
    """
    # 列高
    src_dim = ws.row_dimensions.get(source_row)
    if src_dim and src_dim.height:
        ws.row_dimensions[target_row].height = src_dim.height

    # 儲存格格式（只複製樣式，不建立合併）
    for col_idx in range(1, ws.max_column + 1):
        src = ws.cell(row=source_row, column=col_idx)
        dst = ws.cell(row=target_row, column=col_idx)
        if src.has_style:
            dst.font = _copy_obj(src.font)
            dst.border = _copy_obj(src.border)
            dst.fill = _copy_obj(src.fill)
            dst.number_format = src.number_format
            dst.alignment = _copy_obj(src.alignment)


def _expand_for_trips(oil_ws, travel_ws, new_oil_name, n_oil_rows, n_travel_rows):
    """若行程數超過範本容量，自動插入欄位擴充。
    n_oil_rows  : 出差筆數（油費報告書 1 筆 = 1 行）
    n_travel_rows: 差旅申請單所需總行數（每筆每天 2 行）
    """
    OIL_START = OIL_DATA_START_ROW
    OIL_TOTAL_ROW = _find_total_row(oil_ws) or 30
    oil_capacity = OIL_TOTAL_ROW - OIL_START

    TRAVEL_START = 7
    TRAVEL_TOTAL_ROW = _find_total_row(travel_ws) or 37
    travel_capacity = TRAVEL_TOTAL_ROW - TRAVEL_START

    # ── 擴充油費報告書 ──────────────────────────
    if n_oil_rows > oil_capacity:
        extra = n_oil_rows - oil_capacity
        oil_src = OIL_TOTAL_ROW - 1   # 插入前最後一個資料列
        oil_ws.insert_rows(OIL_TOTAL_ROW, extra)
        for r in range(OIL_TOTAL_ROW, OIL_TOTAL_ROW + extra):
            _copy_row_style(oil_ws, oil_src, r)
            oil_ws[f"E{r}"] = ARROW_SYMBOL
            oil_ws[f"I{r}"] = f"=H{r}*8"
        new_oil_total = OIL_TOTAL_ROW + extra
        last_data = new_oil_total - 1
        oil_ws[f"H{new_oil_total}"] = f"=SUM(H{OIL_START}:H{last_data})"
        oil_ws[f"H{new_oil_total+1}"] = f"=H{new_oil_total}*8"
        oil_ws[f"A{new_oil_total+2}"] = oil_ws[f"A{OIL_TOTAL_ROW+2}"].value
        oil_ws[f"C{new_oil_total+2}"] = oil_ws[f"C{OIL_TOTAL_ROW+2}"].value

    # ── 擴充差旅申請單 ──────────────────────────
    if n_travel_rows > travel_capacity:
        extra = n_travel_rows - travel_capacity
        # 記住來源列（插入前位置不變）
        r1_src = TRAVEL_TOTAL_ROW - 2   # 最後一個 r1 資料列
        r2_src = TRAVEL_TOTAL_ROW - 1   # 最後一個 r2 備註列
        travel_ws.insert_rows(TRAVEL_TOTAL_ROW, extra)

        # ── 修正 openpyxl 的合併儲存格位移 bug ──
        # insert_rows 正確移動儲存格內容，但有時不會自動移動合併範圍。
        # (1) 把 min_row >= TRAVEL_TOTAL_ROW 的合併範圍手動往下移 extra 列
        to_shift = [
            mr for mr in list(travel_ws.merged_cells.ranges)
            if mr.min_row >= TRAVEL_TOTAL_ROW
        ]
        for mr in to_shift:
            min_r, max_r, min_c, max_c = mr.min_row, mr.max_row, mr.min_col, mr.max_col
            # 直接傳物件（非字串），確保 hash 比對命中
            travel_ws.merged_cells.ranges.discard(mr)
            travel_ws.merge_cells(
                start_row=min_r + extra, start_column=min_c,
                end_row=max_r + extra, end_column=max_c,
            )

        # (2) 截斷橫跨插入點（min_row < TRAVEL_TOTAL_ROW < max_row）的合併範圍
        for mr in list(travel_ws.merged_cells.ranges):
            if mr.min_row < TRAVEL_TOTAL_ROW < mr.max_row:
                min_r, max_r = mr.min_row, mr.max_row
                min_c, max_c = mr.min_col, mr.max_col
                travel_ws.unmerge_cells(
                    start_row=min_r, start_column=min_c,
                    end_row=max_r, end_column=max_c,
                )
                original_max_r = TRAVEL_TOTAL_ROW - 1
                if min_r < original_max_r:
                    travel_ws.merge_cells(
                        start_row=min_r, start_column=min_c,
                        end_row=original_max_r, end_column=max_c,
                    )

        # 複製格式到新插入的空白列
        for k in range(extra):
            new_row = TRAVEL_TOTAL_ROW + k
            src_row = r1_src if k % 2 == 0 else r2_src
            _copy_row_style(travel_ws, src_row, new_row)

        new_travel_total = TRAVEL_TOTAL_ROW + extra
        last_data = new_travel_total - 1
        for col in ["G", "H", "I", "J", "K", "O", "P", "Q", "R"]:
            c = travel_ws[f"{col}{new_travel_total}"]
            if not isinstance(c, MergedCell):
                c.value = f"=SUM({col}{TRAVEL_START}:{col}{last_data})"


def generate_excel(trips: list, year: int, month: int, config: dict) -> Path:
    settings = config["settings"]
    roc_year = year - 1911
    mz = _month_zh(month)
    n_trips = len(trips)

    # ── 載入範本 ──────────────────────────────
    wb = load_workbook(str(TEMPLATE_FILE))

    # ── 重新命名工作表 ─────────────────────────
    old_oil = "油費報告書2月"
    old_travel = "差旅申請單2月"
    new_oil = f"油費報告書{mz}"
    new_travel = f"差旅申請單{mz}"

    _rename_sheet_and_update_refs(wb, old_oil, new_oil)
    _rename_sheet_and_update_refs(wb, old_travel, new_travel)

    oil_ws = wb[new_oil]
    travel_ws = wb[new_travel]

    # ── 計算所需行數 ────────────────────────────
    n_oil_rows = n_trips
    n_travel_rows = sum(t["days"] * 2 for t in trips)

    # ── 若行程超過容量，先擴充表格 ────────────
    _expand_for_trips(oil_ws, travel_ws, new_oil, n_oil_rows, n_travel_rows)

    # ── 重新找合計列（可能已移動）──────────────
    oil_total_row = _find_total_row(oil_ws) or 30
    travel_total_row = _find_total_row(travel_ws) or 37
    TRAVEL_START = 7

    # ── 更新油費報告書表頭 ─────────────────────
    oil_ws["A1"] = f"{roc_year}年"
    oil_ws["B1"] = month          # B1 = 出差月份
    receipt_row = oil_total_row + 2
    oil_ws[f"A{receipt_row}"] = roc_year
    oil_ws[f"C{receipt_row}"] = month

    # ── 更新差旅申請單表頭（報表產出日期）────────
    today = date.today()
    today_roc = today.year - 1911
    travel_ws["N2"] = today_roc    # 報表產出年（民國）
    travel_ws["P2"] = today.month  # 報表產出月
    travel_ws["R2"] = today.day    # 報表產出日

    # ── 合併儲存格安全寫入 ─────────────────────
    def safe_set(ws, coord, value):
        cell = ws[coord]
        if not isinstance(cell, MergedCell):
            cell.value = value

    # ── 清空油費報告書資料列 ────────────────────
    for r in range(OIL_DATA_START_ROW, oil_total_row):
        oil_ws[f"A{r}"] = None
        oil_ws[f"B{r}"] = None
        oil_ws[f"C{r}"] = None
        oil_ws[f"E{r}"] = ARROW_SYMBOL
        oil_ws[f"F{r}"] = None
        oil_ws[f"H{r}"] = None
        oil_ws[f"J{r}"] = None
        oil_ws[f"K{r}"] = None

    # ── 清空差旅申請單手動欄位（全範圍）──────────
    for r in range(TRAVEL_START, travel_total_row):
        for col in ["I", "J", "L", "M", "N", "O", "P"]:
            safe_set(travel_ws, f"{col}{r}", None)

    # ── 填入出差資料 ────────────────────────────
    travel_offset = 0
    row_pairs = []   # (r1, r2) for each trip, used to ensure A-column vertical merge

    for i, trip in enumerate(trips):
        oil_row = OIL_DATA_START_ROW + i

        # 油費報告書：日期欄顯示完整區間
        oil_ws[f"A{oil_row}"] = trip["month"]
        if trip["days"] > 1:
            end_d = trip.get("end_day", trip["day"] + trip["days"] - 1)
            oil_ws[f"B{oil_row}"] = f"{trip['day']}~{end_d}"
        else:
            oil_ws[f"B{oil_row}"] = trip["day"]
        oil_ws[f"C{oil_row}"] = settings["start_point"]
        oil_ws[f"E{oil_row}"] = ARROW_SYMBOL
        oil_ws[f"F{oil_row}"] = trip["destination"]
        oil_ws[f"H{oil_row}"] = trip["mileage"]
        oil_ws[f"J{oil_row}"] = trip["work_reason"]
        oil_ws[f"K{oil_row}"] = trip["company"]

        # 差旅申請單：第一天（連結油費報告書）
        r1 = TRAVEL_START + travel_offset
        r2 = r1 + 1
        row_pairs.append((r1, r2))
        safe_set(travel_ws, f"A{r1}", i + 1)
        safe_set(travel_ws, f"B{r1}", f"={new_oil}!J{oil_row}")
        safe_set(travel_ws, f"C{r1}", f"={new_oil}!A{oil_row}")
        safe_set(travel_ws, f"D{r1}", f"={new_oil}!B{oil_row}")
        safe_set(travel_ws, f"E{r1}", f"={new_oil}!C{oil_row}")
        safe_set(travel_ws, f"F{r1}", f"={new_oil}!F{oil_row}")
        safe_set(travel_ws, f"G{r1}", f"={new_oil}!H{oil_row}")
        safe_set(travel_ws, f"H{r1}", f"={new_oil}!I{oil_row}")
        safe_set(travel_ws, f"K{r1}", f"=SUM(M{r1}:M{r2})")
        safe_set(travel_ws, f"P{r1}", trip["meal_allowance"])   # 多日出差顯示總膳雜費
        safe_set(travel_ws, f"Q{r1}", f"=SUM(G{r1}*8,I{r1}:K{r1},P{r1})")
        safe_set(travel_ws, f"B{r2}", f"={new_oil}!K{oil_row}")

        # 差旅申請單：第二天起，每天各 2 行（只填日期與起迄，膳雜費已含在第一天）
        for d in range(1, trip["days"]):
            day_num = trip["day"] + d
            rd1 = r1 + d * 2
            safe_set(travel_ws, f"C{rd1}", trip["month"])
            safe_set(travel_ws, f"D{rd1}", day_num)
            safe_set(travel_ws, f"E{rd1}", settings["start_point"])
            safe_set(travel_ws, f"F{rd1}", trip["destination"])

        travel_offset += trip["days"] * 2

    # ── 確保差旅申請單每筆 A 欄縱向合併（編號跨 r1:r2）──
    existing_merge_starts = {
        (mr.min_row, mr.min_col) for mr in travel_ws.merged_cells.ranges
    }
    for r1, r2 in row_pairs:
        if (r1, 1) not in existing_merge_starts:
            try:
                travel_ws.merge_cells(f"A{r1}:A{r2}")
            except Exception:
                pass

    # ── 儲存 ────────────────────────────────────
    output_name = f"陳峙霖{year}差旅報支申請單(油費連動){mz}.xlsx"
    output_path = BASE_DIR / output_name
    wb.save(str(output_path))
    return output_path


# ════════════════════════════════════════════
# 主程式
# ════════════════════════════════════════════

def main():
    # ── 決定年月 ──────────────────────────────
    today = date.today()

    if len(sys.argv) == 3:
        year, month = int(sys.argv[1]), int(sys.argv[2])
    elif len(sys.argv) == 2:
        year = today.year
        month = int(sys.argv[1])
    else:
        # 預設：上個月
        if today.month == 1:
            year, month = today.year - 1, 12
        else:
            year, month = today.year, today.month - 1

    print(f"\n╔══════════════════════════════════════╗")
    print(f"║  出差申報自動化  {year}年{month}月          ║")
    print(f"╚══════════════════════════════════════╝")

    # ── 載入設定 ──────────────────────────────
    config = load_config()

    # ── Google 行事曆 ─────────────────────────
    print("\n🔐 連接 Google 行事曆（首次執行會彈出瀏覽器授權）...")
    service = get_calendar_service()

    print(f"📅 搜尋關鍵字「{config['settings']['calendar_keyword']}」的行程...")
    events = fetch_events(service, config, year, month)

    if not events:
        print("⚠  本月沒有找到符合條件的行程，結束")
        return

    print(f"   找到 {len(events)} 筆行程，開始解析...")

    # ── 解析事件 ──────────────────────────────
    trips = []
    skipped = []
    for ev in events:
        parsed = parse_event(ev, config)
        if parsed:
            trips.extend(parsed)
        else:
            skipped.append(ev.get("summary", "（無標題）"))

    if skipped:
        print(f"   ⚠  以下 {len(skipped)} 筆無法解析日期，已略過：")
        for s in skipped:
            print(f"      • {s}")

    if not trips:
        print("⚠  沒有可用的出差記錄，結束")
        return

    # ── 同日合併 ──────────────────────────────
    trips = merge_same_day_trips(trips, config)

    # ── 顯示解析結果 ──────────────────────────
    print(f"\n{'─'*62}")
    print(f"  共解析出 {len(trips)} 筆出差記錄：")
    print(f"{'─'*62}")
    for i, trip in enumerate(trips, 1):
        if trip["days"] > 1:
            end_d = trip.get("end_day", trip["day"] + trip["days"] - 1)
            date_str = f"{trip['month']:02d}/{trip['day']:02d}~{end_d:02d}（{trip['days']}天）"
        else:
            date_str = f"{trip['month']:02d}/{trip['day']:02d}"
        print(f"  [{i:2d}]  {date_str}  {trip['destination'] or '⚠未知'}  ─  {trip['company']}")
        print(f"        里程 {trip['mileage']} km  ｜  事由：{trip['work_reason']}  ｜  膳雜費：{trip['meal_allowance']} 元")

    confirmed = trips

    # ── 產生 Excel ────────────────────────────
    print(f"\n📊 產生報表中（共 {len(confirmed)} 筆）...")
    output_path = generate_excel(confirmed, year, month, config)

    print(f"\n✅ 已產生：{output_path.name}")
    print("\n─── 請開啟 Excel 手動補填以下欄位 ───────────")
    print("   差旅申請單  I 欄：火車費")
    print("   差旅申請單  J 欄：高鐵費")
    print("   差旅申請單  L 欄：其他費用說明（計程車/停車費）")
    print("   差旅申請單  M 欄：其他費用金額")
    print("   差旅申請單  O 欄：旅館費（如有住宿）")
    print("   差旅申請單  R 欄：審核金額（主管填寫）")
    print("─────────────────────────────────────────────")

    if sys.platform == "darwin":
        os.system(f'open "{output_path}"')

    # ── 寄送 Email 通知 ───────────────────────
    _send_email_via_mail_app(output_path, year, month, confirmed)


def _send_email_via_mail_app(output_path: Path, year: int, month: int, trips: list):
    """
    用 Gmail SMTP 寄送報表通知。
    密碼從環境變數 GMAIL_APP_PASSWORD 或 macOS Keychain 讀取。
    """
    import smtplib
    import subprocess
    from email.message import EmailMessage

    SENDER   = "foren0516@gmail.com"
    RECIPIENT = "foren@cc-sustain.com"
    KEYCHAIN_SERVICE = "businesstrip_gmail"

    # ── 取得 Gmail App 密碼（環境變數優先，其次 macOS Keychain）──
    password = os.environ.get("GMAIL_APP_PASSWORD")
    if not password and sys.platform == "darwin":
        try:
            result = subprocess.run(
                ["security", "find-generic-password",
                 "-a", SENDER, "-s", KEYCHAIN_SERVICE, "-w"],
                capture_output=True, text=True, check=True
            )
            password = result.stdout.strip()
        except subprocess.CalledProcessError:
            pass
    if not password:
        print("\n⚠  Email 未寄出：找不到 Gmail App 密碼。")
        print("   請設定環境變數 GMAIL_APP_PASSWORD，或在 macOS 執行：")
        print(f'   security add-generic-password -a {SENDER} -s {KEYCHAIN_SERVICE} -w <App密碼>')
        return

    # ── 組裝信件（使用 EmailMessage，原生支援 UTF-8）──
    subject = f"【自動通知】{year}年{month}月 差旅報支申請單 已產生"
    body = f"{year}年{month}月份的差旅報支申請單已經自動產生，請參閱附件。"

    msg = EmailMessage()
    msg["From"] = SENDER
    msg["To"] = RECIPIENT
    msg["Subject"] = subject
    msg.set_content(body)

    with open(output_path, "rb") as f:
        msg.add_attachment(f.read(),
                           maintype="application",
                           subtype="octet-stream",
                           filename=output_path.name)

    # ── 透過 Gmail SMTP 寄出 ──
    try:
        with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.login(SENDER, password)
            smtp.send_message(msg)
        print(f"\n📧 報表已寄至 {RECIPIENT}（寄件人：{SENDER}）")
    except Exception as e:
        print(f"\n⚠  Email 寄送失敗：{e}")


if __name__ == "__main__":
    main()
